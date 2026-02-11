"""
Screener Quarterly Fetcher (Production-Safe Version)
----------------------------------------------------
- Recency-based selection (standalone vs consolidated)
- Rate-limit handling (429 retry with backoff)
- Local HTML caching (avoids repeated hits)
- YoY for last 3 quarters
- Per-quarter + stock summary outputs
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
import sys
import random

BASE_URL_CONSOL = "https://www.screener.in/company/{symbol}/consolidated/"
BASE_URL_STANDALONE = "https://www.screener.in/company/{symbol}/"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36"
}

REQUEST_DELAY = 2.5  # safer delay
MAX_RETRIES = 5
CACHE_DIR = "screener_cache"
os.makedirs(CACHE_DIR, exist_ok=True)


# ------------------------------------------------
# Safe Request with Retry + Backoff
# ------------------------------------------------

def safe_request(url):
    for attempt in range(MAX_RETRIES):
        response = requests.get(url, headers=HEADERS, timeout=20)

        if response.status_code == 429:
            wait = 5 + random.uniform(0, 3)
            print(f"⚠ Rate limited. Sleeping {wait:.1f}s...")
            time.sleep(wait)
            continue

        response.raise_for_status()
        return response

    raise Exception(f"Max retries exceeded for {url}")


# ------------------------------------------------
# HTML Fetch with Caching
# ------------------------------------------------

def fetch_html(symbol, consolidated=False):
    suffix = "_consol" if consolidated else "_standalone"
    cache_file = os.path.join(CACHE_DIR, f"{symbol}{suffix}.html")

    if os.path.exists(cache_file):
        with open(cache_file, "r", encoding="utf-8") as f:
            return f.read()

    url = (BASE_URL_CONSOL if consolidated else BASE_URL_STANDALONE).format(symbol=symbol)
    response = safe_request(url)

    with open(cache_file, "w", encoding="utf-8") as f:
        f.write(response.text)

    time.sleep(REQUEST_DELAY)
    return response.text


# ------------------------------------------------
# Parse Quarterly Section
# ------------------------------------------------

def parse_quarterly(html):
    soup = BeautifulSoup(html, "html.parser")
    section = soup.find("section", id="quarters")

    if section is None:
        return None

    table = section.find("table")
    rows = table.find_all("tr")

    headers = [th.text.strip() for th in rows[0].find_all("th")]
    data = []

    for row in rows[1:]:
        cols = [td.text.strip().replace(',', '') for td in row.find_all("td")]
        if cols:
            data.append(cols)

    if not data:
        return None

    df = pd.DataFrame(data, columns=headers)
    df.set_index(df.columns[0], inplace=True)
    df = df.T

    rename_map = {}
    for col in df.columns:
        c = col.lower()
        if 'sale' in c or 'revenue' in c:
            rename_map[col] = 'Sales'
        elif 'net profit' in c or 'profit after tax' in c or c == 'profit':
            rename_map[col] = 'Net Profit'

    df.rename(columns=rename_map, inplace=True)

    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df.index.name = 'Quarter'
    df.reset_index(inplace=True)

    # Remove TTM before parsing dates
    df = df[df['Quarter'] != 'TTM'].copy()

    df['__order'] = pd.to_datetime(df['Quarter'], format='%b %Y', errors='coerce')
    df.sort_values('__order', inplace=True)
    df.drop(columns='__order', inplace=True)

    return df


# ------------------------------------------------
# Recency-Based Selection
# ------------------------------------------------

def fetch_quarterly_results(symbol):
    html_consol = fetch_html(symbol, consolidated=True)
    html_standalone = fetch_html(symbol, consolidated=False)

    df_consol = parse_quarterly(html_consol)
    df_standalone = parse_quarterly(html_standalone)

    def latest(df):
        if df is None or df.empty:
            return None
        return pd.to_datetime(df['Quarter'], format='%b %Y', errors='coerce').max()

    date_consol = latest(df_consol)
    date_standalone = latest(df_standalone)

    if date_consol is None and date_standalone is None:
        raise ValueError("No quarterly data found")

    if date_standalone is None:
        return df_consol
    if date_consol is None:
        return df_standalone

    return df_standalone if date_standalone > date_consol else df_consol


# ------------------------------------------------
# YoY Logic
# ------------------------------------------------

def analyze_stock_for_quarter(df, target_idx, symbol):
    curr = df.iloc[target_idx]
    prev = df.iloc[target_idx - 4] if target_idx - 4 >= 0 else None

    sales_yoy = None
    profit_yoy = None
    yoy_available = prev is not None

    if prev is not None:
        if pd.notna(prev['Sales']) and prev['Sales'] != 0:
            sales_yoy = round((curr['Sales'] - prev['Sales']) / abs(prev['Sales']) * 100, 2)

        if pd.notna(prev['Net Profit']) and prev['Net Profit'] != 0:
            profit_yoy = round((curr['Net Profit'] - prev['Net Profit']) / abs(prev['Net Profit']) * 100, 2)

    return {
        'Symbol': symbol,
        'Quarter': curr['Quarter'],
        'Sales (₹ Cr)': curr['Sales'],
        'Profit (₹ Cr)': curr['Net Profit'],
        'YoY Quarter': prev['Quarter'] if prev is not None else None,
        'YoY Sales (₹ Cr)': prev['Sales'] if prev is not None else None,
        'YoY Profit (₹ Cr)': prev['Net Profit'] if prev is not None else None,
        'Sales YoY (%)': sales_yoy,
        'Profit YoY (%)': profit_yoy,
        'YoY Available': yoy_available,
    }


def analyze_stock(symbol):
    df = fetch_quarterly_results(symbol)

    if df is None or len(df) < 5:
        return []

    results = []
    for offset in [-1, -2, -3]:
        idx = len(df) + offset
        if idx < 0:
            continue
        row = analyze_stock_for_quarter(df, idx, symbol)
        results.append(row)

    return results


def analyze_ticker_file(ticker_file):
    all_rows = []

    with open(ticker_file) as f:
        symbols = [line.strip() for line in f if line.strip()]

    for i, symbol in enumerate(symbols, 1):
        try:
            print(f"[{i}/{len(symbols)}] Processing {symbol}")
            rows = analyze_stock(symbol)
            all_rows.extend(rows)
        except Exception as e:
            print(f"❌ {symbol}: {e}")

    df = pd.DataFrame(all_rows)
    if df.empty:
        return df

    cols = ['Symbol'] + [c for c in df.columns if c != 'Symbol']
    return df[cols]


# ------------------------------------------------
# Main (Dashboard + Conditional Formatting)
# ------------------------------------------------

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

if __name__ == '__main__':
    if len(sys.argv) < 2:
        raise ValueError("Usage: python screener_quarterly_fetcher.py <ticker_file.txt>")

    ticker_file = sys.argv[1]
    df = analyze_ticker_file(ticker_file)

    if df.empty:
        print("No data generated")
        sys.exit(0)

    # ---------------- Dashboard Construction ----------------
    dashboard_rows = []

    for symbol, g in df.groupby('Symbol'):
        # Proper chronological sorting (NOT string sort)
        g = g.copy()
        g['__order'] = pd.to_datetime(g['Quarter'], format='%b %Y', errors='coerce')
        g = g.sort_values('__order')
        g = g.drop(columns='__order')

        sales_list = g['Sales YoY (%)'].tolist()
        profit_list = g['Profit YoY (%)'].tolist()

        consistency = sum(1 for x in profit_list if x is not None and x > 0)

        acceleration = None
        if len(profit_list) >= 3 and None not in profit_list[-3:]:
            if profit_list[-1] > profit_list[-2] > profit_list[-3]:
                acceleration = "Up"
            elif profit_list[-1] < profit_list[-2] < profit_list[-3]:
                acceleration = "Down"
            else:
                acceleration = "Flat"

        score = consistency * 2
        if acceleration == "Up":
            score += 2
        elif acceleration == "Down":
            score -= 1

        row = {
            'Symbol': symbol,
            'Sales Q1': sales_list[-3] if len(sales_list) >= 3 else None,
            'Sales Q2': sales_list[-2] if len(sales_list) >= 2 else None,
            'Sales Q3': sales_list[-1] if len(sales_list) >= 1 else None,
            'Profit Q1': profit_list[-3] if len(profit_list) >= 3 else None,
            'Profit Q2': profit_list[-2] if len(profit_list) >= 2 else None,
            'Profit Q3': profit_list[-1] if len(profit_list) >= 1 else None,
            'Consistency': consistency,
            'Acceleration': acceleration,
            'Score': score
        }
        dashboard_rows.append(row)

    dashboard_df = pd.DataFrame(dashboard_rows)
    dashboard_df.sort_values('Score', ascending=False, inplace=True)

    output_file = "quarterly_outputs/dashboard.xlsx"
    dashboard_df.to_excel(output_file, index=False, sheet_name="Dashboard")

    # ---------------- Conditional Formatting ----------------
    wb = load_workbook(output_file)
    ws = wb.active

    color_rule = ColorScaleRule(
        start_type='num', start_value=-20, start_color='F8696B',
        mid_type='num', mid_value=0, mid_color='FFEB84',
        end_type='num', end_value=30, end_color='63BE7B'
    )

    ws.conditional_formatting.add(
        f"B2:G{ws.max_row}",
        color_rule
    )

    wb.save(output_file)

    print("✅ Dashboard generated with visual heatmap formatting")
